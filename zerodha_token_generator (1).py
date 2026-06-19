
#!/usr/bin/env python3
"""
Zerodha Kite Connect  ─  Token Generator + Config Updater
══════════════════════════════════════════════════════════════════════════════
Single daily-run script that:

  STEP 1  Generate (or reuse) today's access token via Kite Connect login.
  STEP 2  Save token to zerodha_token.json  (same folder as this script).
  STEP 3  Write api_key / api_secret / access_token → zerodha_config.xlsx
            B4  api_key  |  B5  api_secret  |  B6  access_token
  STEP 4  Fetch all upcoming expiry dates from Zerodha (NFO + BFO) and
          install live dropdown lists in zerodha_config.xlsx:
            B10  NIFTY  |  B11  BANKNIFTY  |  B12  SENSEX

  If zerodha_config.xlsx does NOT exist it is created automatically
  in the same folder as this script, with the standard layout and
  formatting (matching the reference screenshot).

  All other cells, formatting, expiry ranges, strike ranges, and notes
  in an existing zerodha_config.xlsx are left completely untouched.

Requirements:
    pip install kiteconnect openpyxl
══════════════════════════════════════════════════════════════════════════════
"""

import hashlib
import json
import os
import webbrowser
import datetime as _dt

# ╔══════════════════════════════════════════════════════════════╗
#  CONFIG — Edit only this section
# ╚══════════════════════════════════════════════════════════════╝
API_KEY    = "paste here zerodha api key"
API_SECRET = "paste here zerodha api secrate"
# ╚══════════════════════════════════════════════════════════════╝

# All files live next to this .py script — no hardcoding needed
_HERE       = os.path.dirname(os.path.abspath(__file__))
CONFIG_XLSX = os.path.join(_HERE, "zerodha_config.xlsx")
TOKEN_FILE  = os.path.join(_HERE, "zerodha_token.json")

# Sheet / cell constants
CONFIG_SHEET   = "Config"
HELPER_SHEET   = "_ExpiryLists"
CELL_NIFTY     = "B10"
CELL_BANKNIFTY = "B11"
CELL_SENSEX    = "B12"
COL_NIFTY      = 1
COL_BANKNIFTY  = 2
COL_SENSEX     = 3


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 0 — Create zerodha_config.xlsx from scratch if it doesn't exist
# ══════════════════════════════════════════════════════════════════════════════

def create_config_xlsx() -> None:
    """
    Build a brand-new zerodha_config.xlsx matching the standard layout:

    Row 1  : Title banner
    Row 3  : Section header — API Credentials
    Row 4  : api_key        (A4 label | B4 value)
    Row 5  : api_secret     (A5 label | B5 value)
    Row 6  : access_token   (A6 label | B6 value)
    Row 8  : Section header — Expiry Dates
    Row 9  : Column headers (Index | Expiry Date)
    Row 10 : NIFTY     expiry
    Row 11 : BANKNIFTY expiry
    Row 12 : SENSEX    expiry
    Row 14 : Section header — Strike Range
    Row 15 : Column headers (Index | Range pts)
    Row 16 : NIFTY     2500
    Row 17 : BANKNIFTY 2500
    Row 18 : SENSEX    1000
    Row 20 : Section header — Notes
    Row 21-23: Note bullets
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n⚠️  openpyxl not installed — cannot create config file.")
        print("    Install with:  pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = CONFIG_SHEET

    # ── Colour palette (matching screenshot) ──────────────────────────────────
    DARK_NAVY   = "1F3864"   # title / section headers
    MID_BLUE    = "2E75B6"   # sub-headers (Index / Expiry Date row)
    DARK_GREEN  = "375623"   # section header — Strike Range
    PURPLE      = "7030A0"   # section header — Notes
    GREEN_TEXT  = "00B050"   # cell values (credentials + expiry dates)
    WHITE       = "FFFFFF"
    LIGHT_GRAY  = "F2F2F2"   # alternating row tint

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def white_bold(size=11):
        return Font(name="Arial", bold=True, color=WHITE, size=size)

    def green_val(size=11):
        return Font(name="Arial", color=GREEN_TEXT, size=size)

    def col_hdr_font():
        return Font(name="Arial", bold=True, color=WHITE, size=10)

    thin = Side(style="thin", color="CCCCCC")
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 5   # visual padding

    # ── Row heights ───────────────────────────────────────────────────────────
    for r in range(1, 24):
        ws.row_dimensions[r].height = 18

    # helper: merge + style a full-width header cell
    def section_header(row, text, fill_color, font=None):
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value     = f"  {text}"
        cell.fill      = hdr_fill(fill_color)
        cell.font      = font or white_bold(11)
        cell.alignment = left
        cell.border    = box

    # ── Row 1 : Title ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value     = "  Zerodha API Configuration  \u2500  zerodha_config.xlsx"
    t.fill      = hdr_fill(DARK_NAVY)
    t.font      = Font(name="Arial", bold=True, color=WHITE, size=13)
    t.alignment = left
    t.border    = box

    # ── Row 3 : API Credentials header ───────────────────────────────────────
    section_header(3, "API Credentials", DARK_NAVY)

    # ── Rows 4-6 : Credentials ────────────────────────────────────────────────
    cred_rows = {4: "api_key", 5: "api_secret", 6: "access_token"}
    for row, label in cred_rows.items():
        a = ws[f"A{row}"]
        b = ws[f"B{row}"]
        a.value     = label
        a.font      = Font(name="Arial", bold=True, size=10)
        a.alignment = left
        a.border    = box
        b.value     = ""          # filled at runtime by update_config_credentials
        b.font      = green_val()
        b.alignment = left
        b.border    = box
        if row % 2 == 0:
            a.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # ── Row 8 : Expiry Dates header ───────────────────────────────────────────
    section_header(8, "Expiry Dates  (YYYY-MM-DD)", DARK_NAVY)

    # ── Row 9 : Column sub-headers ────────────────────────────────────────────
    for col, txt in [("A", "Index"), ("B", "Expiry Date")]:
        c = ws[f"{col}9"]
        c.value     = txt
        c.fill      = hdr_fill(MID_BLUE)
        c.font      = col_hdr_font()
        c.alignment = center
        c.border    = box

    # ── Rows 10-12 : Index rows ───────────────────────────────────────────────
    expiry_defaults = {10: ("NIFTY", ""), 11: ("BANKNIFTY", ""), 12: ("SENSEX", "")}
    for row, (idx, default_date) in expiry_defaults.items():
        a, b = ws[f"A{row}"], ws[f"B{row}"]
        a.value     = idx
        a.font      = Font(name="Arial", bold=True, size=10)
        a.alignment = left
        a.border    = box
        b.value     = default_date
        b.font      = green_val()
        b.alignment = left
        b.border    = box
        if row % 2 == 0:
            a.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            b.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # ── Row 14 : Strike Range header ──────────────────────────────────────────
    section_header(14, "Strike Range  (\u00b1 points from ATM)", PURPLE)

    # ── Row 15 : Column sub-headers ───────────────────────────────────────────
    for col, txt in [("A", "Index"), ("B", "Range (pts)")]:
        c = ws[f"{col}15"]
        c.value     = txt
        c.fill      = hdr_fill(MID_BLUE)
        c.font      = col_hdr_font()
        c.alignment = center
        c.border    = box

    # ── Rows 16-18 : Strike ranges ────────────────────────────────────────────
    strike_rows = {16: ("NIFTY", 2500), 17: ("BANKNIFTY", 2500), 18: ("SENSEX", 1000)}
    for row, (idx, pts) in strike_rows.items():
        a, b = ws[f"A{row}"], ws[f"B{row}"]
        a.value     = idx
        a.font      = Font(name="Arial", bold=True, size=10)
        a.alignment = left
        a.border    = box
        b.value     = pts
        b.font      = green_val()
        b.alignment = left
        b.border    = box
        if row % 2 == 0:
            a.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            b.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # ── Row 20 : Notes header ─────────────────────────────────────────────────
    section_header(20, "Notes", DARK_NAVY)

    # ── Rows 21-23 : Note bullets ─────────────────────────────────────────────
    notes = [
        "  \u2022 Run this script once per day (after market login) to refresh access_token.",
        "  \u2022 Expiry dropdowns in B10, B11, B12 are auto-updated from live Zerodha feed.",
        "  \u2022 Other scripts read this file automatically from the same folder.",
    ]
    for row, note in enumerate(notes, 21):
        ws.merge_cells(f"A{row}:B{row}")
        c = ws[f"A{row}"]
        c.value     = note
        c.font      = Font(name="Arial", size=9, italic=True)
        c.alignment = left
        c.border    = box

    wb.save(CONFIG_XLSX)
    print(f"\n✅  zerodha_config.xlsx created  →  {CONFIG_XLSX}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — Token generation helpers
# ══════════════════════════════════════════════════════════════════════════════

def generate_checksum(api_key, request_token, api_secret):
    raw = f"{api_key}{request_token}{api_secret}"
    return hashlib.sha256(raw.encode()).hexdigest()


def get_login_url(api_key):
    return f"https://kite.zerodha.com/connect/login?v=3&api_key={api_key}"


def exchange_token(api_key, request_token, api_secret):
    import urllib.request, urllib.parse
    checksum = generate_checksum(api_key, request_token, api_secret)
    payload  = urllib.parse.urlencode({
        "api_key": api_key, "request_token": request_token, "checksum": checksum,
    }).encode()
    req = urllib.request.Request(
        "https://api.kite.trade/session/token", data=payload, method="POST",
        headers={"X-Kite-Version": "3",
                 "Content-Type": "application/x-www-form-urlencoded"},
    )
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode())


def save_token(data, path):
    out = {
        "access_token": data.get("access_token"),
        "public_token": data.get("public_token"),
        "user_id":      data.get("user_id"),
        "user_name":    data.get("user_name"),
        "generated_at": _dt.datetime.now().isoformat(),
    }
    with open(path, "w") as f:
        json.dump(out, f, indent=2)
    print(f"\n✅  Token saved  →  {path}")


def load_saved_token(path):
    if not os.path.exists(path):
        return None
    with open(path) as f:
        data = json.load(f)
    gen = _dt.datetime.fromisoformat(data.get("generated_at", "2000-01-01"))
    return data if gen.date() == _dt.datetime.today().date() else None


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — Write credentials to Config sheet  (B4 / B5 / B6)
# ══════════════════════════════════════════════════════════════════════════════

def update_config_credentials(api_key, api_secret, access_token):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("\n⚠️  openpyxl not found.  Install with:  pip install openpyxl")
        return

    try:
        wb = load_workbook(CONFIG_XLSX)
        ws = wb[CONFIG_SHEET]
        ws["B4"] = api_key
        ws["B5"] = api_secret
        ws["B6"] = access_token
        wb.save(CONFIG_XLSX)
        print(f"\n✅  Credentials written  →  {CONFIG_XLSX}  (B4 / B5 / B6)")
        print(f"    api_key      : {api_key}")
        print(f"    api_secret   : {api_secret}")
        print(f"    access_token : {access_token[:10]}…{access_token[-6:]}")
    except Exception as e:
        print(f"\n❌  Credential update failed: {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — Fetch expiries + update dropdowns  (B10 / B11 / B12)
# ══════════════════════════════════════════════════════════════════════════════

def fetch_expiries(api_key, access_token):
    try:
        from kiteconnect import KiteConnect
    except ImportError:
        print("\n⚠️  kiteconnect not installed.  Install with:  pip install kiteconnect")
        return {}

    kite  = KiteConnect(api_key=api_key)
    kite.set_access_token(access_token)
    today  = _dt.date.today()
    result = {"NIFTY": [], "BANKNIFTY": [], "SENSEX": []}

    print("\n  Downloading NFO instruments ...")
    try:
        nfo = kite.instruments("NFO")
    except Exception as e:
        print(f"  ❌  NFO fetch failed: {e}")
        nfo = []

    for idx in ("NIFTY", "BANKNIFTY"):
        dates = sorted({
            i["expiry"] for i in nfo
            if i["name"] == idx
            and i["instrument_type"] in ("CE", "PE")
            and isinstance(i["expiry"], _dt.date)
            and i["expiry"] >= today
        })
        result[idx] = [d.strftime("%Y-%m-%d") for d in dates]
        print(f"  ✓ {idx:<12}  {len(result[idx])} expiries  "
              f"({result[idx][0] if result[idx] else '—'}  →  "
              f"{result[idx][-1] if result[idx] else '—'})")

    print("  Downloading BFO instruments ...")
    try:
        bfo = kite.instruments("BFO")
    except Exception as e:
        print(f"  ❌  BFO fetch failed: {e}")
        bfo = []

    dates = sorted({
        i["expiry"] for i in bfo
        if i["name"] == "SENSEX"
        and i["instrument_type"] in ("CE", "PE")
        and isinstance(i["expiry"], _dt.date)
        and i["expiry"] >= today
    })
    result["SENSEX"] = [d.strftime("%Y-%m-%d") for d in dates]
    print(f"  ✓ {'SENSEX':<12}  {len(result['SENSEX'])} expiries  "
          f"({result['SENSEX'][0] if result['SENSEX'] else '—'}  →  "
          f"{result['SENSEX'][-1] if result['SENSEX'] else '—'})")

    return result


def update_config_expiry_dropdowns(expiries):
    if not expiries:
        print("\n⚠️  No expiry data — skipping dropdown update.")
        return

    try:
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n⚠️  openpyxl not installed — skipping dropdown update.")
        return

    try:
        wb     = load_workbook(CONFIG_XLSX)
        cfg_ws = wb[CONFIG_SHEET]

        # Rebuild _ExpiryLists hidden sheet
        if HELPER_SHEET in wb.sheetnames:
            del wb[HELPER_SHEET]
        helper             = wb.create_sheet(HELPER_SHEET)
        helper.sheet_state = "hidden"

        index_cols = [
            ("NIFTY",     COL_NIFTY,     expiries.get("NIFTY", [])),
            ("BANKNIFTY", COL_BANKNIFTY, expiries.get("BANKNIFTY", [])),
            ("SENSEX",    COL_SENSEX,    expiries.get("SENSEX", [])),
        ]
        range_refs = {}

        for idx, col, dates in index_cols:
            helper.cell(1, col, idx)
            for ri, d in enumerate(dates, 2):
                helper.cell(ri, col, d)
            col_letter    = get_column_letter(col)
            range_refs[idx] = (
                f"'{HELPER_SHEET}'!${col_letter}$2:${col_letter}${1 + len(dates)}"
            ) if dates else None

        # Remove old DataValidations on B10/B11/B12
        cells_to_clear = {CELL_NIFTY, CELL_BANKNIFTY, CELL_SENSEX}
        cfg_ws.data_validations.dataValidation = [
            dv for dv in cfg_ws.data_validations.dataValidation
            if not any(c in str(dv.sqref) for c in cells_to_clear)
        ]

        # Attach fresh dropdowns
        dropdown_map = {
            CELL_NIFTY:     ("NIFTY",     range_refs.get("NIFTY")),
            CELL_BANKNIFTY: ("BANKNIFTY", range_refs.get("BANKNIFTY")),
            CELL_SENSEX:    ("SENSEX",    range_refs.get("SENSEX")),
        }
        for cell_addr, (idx, ref) in dropdown_map.items():
            if not ref:
                print(f"  [warn] No expiry dates for {idx} — dropdown skipped.")
                continue
            dv = DataValidation(
                type="list", formula1=ref, allow_blank=True, showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid date",
                error=f"Please choose a date from the {idx} expiry list.",
                showInputMessage=True,
                promptTitle=f"{idx} Expiry",
                prompt=f"Select the expiry date for {idx} from the dropdown.",
            )
            dv.add(cell_addr)
            cfg_ws.add_data_validation(dv)

        # Keep valid existing value; else default to nearest expiry
        for cell_addr, (idx, _) in dropdown_map.items():
            valid_dates = expiries.get(idx, [])
            if not valid_dates:
                continue
            cell = cfg_ws[cell_addr]
            if str(cell.value or "").strip() not in valid_dates:
                cell.value = valid_dates[0]
                print(f"  ↻  {idx} expiry set to nearest: {valid_dates[0]}")

        wb.save(CONFIG_XLSX)
        print(f"\n✅  Expiry dropdowns written  →  {CONFIG_XLSX}  (B10 / B11 / B12)")

    except Exception as e:
        print(f"\n❌  Dropdown update failed: {e}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 4 — Main entry point
# ══════════════════════════════════════════════════════════════════════════════

def banner():
    print("=" * 62)
    print("  🟢  Zerodha Kite Connect  ─  Token Generator + Config Updater")
    print("=" * 62)
    print(f"  Script folder : {_HERE}")
    print(f"  Config file   : {CONFIG_XLSX}")
    print(f"  Token file    : {TOKEN_FILE}")
    print("=" * 62)


def main():
    banner()

    if "your_api_key" in API_KEY or "your_api_secret" in API_SECRET:
        print("\n⚠️  Please set API_KEY and API_SECRET at the top of this script.\n")
        return

    # ── Ensure zerodha_config.xlsx exists ────────────────────────
    if not os.path.exists(CONFIG_XLSX):
        print(f"\n📄  zerodha_config.xlsx not found — creating it now ...")
        create_config_xlsx()

    access_token = None

    # ── STEP 1 : Reuse today's cached token if available ─────────
    cached = load_saved_token(TOKEN_FILE)
    if cached:
        print(f"\n✅  Found a valid token for today ({_dt.datetime.today().date()})")
        print(f"    User  : {cached.get('user_name')} ({cached.get('user_id')})")
        print(f"    Token : {cached.get('access_token')[:10]}…")
        if input("\n    Use this cached token? [Y/n]: ").strip().lower() != "n":
            print("\n    ✔ Using cached access token.")
            access_token = cached.get("access_token")

    # ── STEP 1 (fresh) : Browser login + token exchange ──────────
    if access_token is None:
        login_url = get_login_url(API_KEY)
        print(f"\n📌  Step 1 — Login to Zerodha")
        print(f"    Opening browser: {login_url}")
        webbrowser.open(login_url)

        print(f"\n📌  Step 2 — After login, copy the 'request_token' from the redirect URL.")
        print(f"    URL looks like:")
        print(f"    https://your-redirect-url.com/?request_token=XXXXXX&action=login&status=success")
        print()
        request_token = input("    Paste your request_token here: ").strip()
        if not request_token:
            print("\n❌  No request token provided. Exiting.")
            return

        print(f"\n📌  Step 3 — Exchanging for access token…")
        try:
            resp = exchange_token(API_KEY, request_token, API_SECRET)
        except Exception as e:
            print(f"\n❌  Token exchange failed: {e}")
            return

        if resp.get("status") != "success":
            print(f"\n❌  API error: {resp.get('message', 'Unknown error')}")
            return

        data         = resp.get("data", {})
        access_token = data.get("access_token")
        print(f"\n🎉  Access token generated!")
        print(f"    User  : {data.get('user_name')} ({data.get('user_id')})")
        print(f"    Token : {access_token[:10]}…{access_token[-6:]}")

        # STEP 2 — Save token JSON
        save_token(data, TOKEN_FILE)

    # ── STEP 3 : Write credentials to Config sheet ───────────────
    print("\n" + "─" * 62)
    print("  Updating credentials in zerodha_config.xlsx …")
    update_config_credentials(API_KEY, API_SECRET, access_token)

    # ── STEP 4 : Fetch expiries + update dropdowns ───────────────
    print("\n" + "─" * 62)
    print("  Fetching live expiry dates from Zerodha API …")
    expiries = fetch_expiries(API_KEY, access_token)
    update_config_expiry_dropdowns(expiries)

    print("\n" + "═" * 62)
    print("  ✅  All done!  zerodha_config.xlsx is fully up to date.")
    print("═" * 62)
    print(f"""
  Usage in your trading script:

    import json
    from kiteconnect import KiteConnect

    with open(r"{TOKEN_FILE}") as f:
        token_data = json.load(f)

    kite = KiteConnect(api_key="{API_KEY}")
    kite.set_access_token(token_data["access_token"])

    # ✅ Ready to trade!
    print(kite.profile())
""")


if __name__ == "__main__":
    main()
