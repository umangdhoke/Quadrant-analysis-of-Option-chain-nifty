"""
oi_recorder_pro.py
═══════════════════════════════════════════════════════════════════════════════
Professional OI Analysis Recorder  —  5-Second Tick Logger
Zerodha KiteTicker WebSocket  |  xlwings COM

SHEETS
──────
  "OI Log"    — Full 16-column time-series log, one row every 5 seconds
  "OI Chart"  — 3 embedded live charts (no timeframe dropdown)
                 Chart 1 : OI Quadrant Trend (4 lines)
                 Chart 2 : PCR + Net Bias Trend (signal zones)
                 Chart 3 : OI Build Rate ΔCE / ΔPE per cycle

DERIVED METRICS (every 5 sec)
──────────────────────────────
  CE ITM / CE OTM / PE OTM / PE ITM   — raw quadrant OI (spot-boundary, no drops)
  Total CE OI  = CE ITM + CE OTM      — cross-check: equals sum of all CE strikes
  Total PE OI  = PE OTM + PE ITM      — cross-check: equals sum of all PE strikes
  PCR          = Total PE / Total CE   — core directional indicator
  Net Bias     = CE_OTM − PE_OTM      — writing sentiment above vs below spot
  OI Imbalance = CE_ITM − PE_ITM      — demand-side directional pressure
  ΔCE OI       = change in total CE OI vs prev row
  ΔPE OI       = change in total PE OI vs prev row
  ΔPCR         = PCR change vs prev row
  Δspot        = price move vs prev row
  Signal       = STRONG BUY / BUY / BULL CONFIRM / NEUTRAL /
                 BEAR CONFIRM / SELL / STRONG SELL / COVER RALLY / UNWIND SELL

ALGO SIGNAL RULES
──────────────────
  PCR > 1.3 & ΔCE<0 & Δspot>0  → STRONG BUY    (CE unwinding, mkt rising)
  PCR < 0.7 & ΔPE<0 & Δspot<0  → STRONG SELL   (PE unwinding, mkt falling)
  CE_OTM surging & Δspot>0      → SELL           (resistance wall building)
  PE_OTM surging & Δspot<0      → BUY            (support floor building)
  ΔCE>0 & Δspot>0               → BULL CONFIRM   (long buildup)
  ΔCE>0 & Δspot<0               → BEAR CONFIRM   (short buildup)
  ΔCE<0 & Δspot>0               → COVER RALLY    (short covering)
  ΔPE<0 & Δspot<0               → UNWIND SELL    (long unwinding)

USAGE
─────
    python oi_recorder_pro.py   (zerodha_config.xlsx must be in same folder)
    Ctrl+C → graceful stop, file saved, Excel stays open.

DEPENDENCIES
────────────
    pip install kiteconnect xlwings openpyxl
═══════════════════════════════════════════════════════════════════════════════
"""

import sys, os, time, datetime, math, threading, signal, collections, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from kiteconnect import KiteConnect, KiteTicker
import xlwings as xw


# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
CONFIG_FILE  = "zerodha_config.xlsx"
OUTPUT_FILE  = os.path.abspath("oi_analysis_pro_" + datetime.datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx")
CSV_BACKUP_FILE = os.path.splitext(OUTPUT_FILE)[0] + "_backup.csv"   # ← crash-safe backup
AUTO_SAVE_CYCLES = 60   # save Excel to disk every 60 cycles (~5 min)
RECORD_SEC   = 5

SHEET_LOG    = "OI Log"
SHEET_CHART  = "OI Chart"

# PCR threshold bands for signal generation and chart reference lines
PCR_OVERSOLD  = 1.3    # above this → market oversold puts → bullish fade
PCR_OVERBOUGHT = 0.7   # below this → market oversold calls → bearish fade

# OI surge threshold: % change per 5-sec cycle that is considered "sharp"
OI_SURGE_PCT  = 0.35   # 0.35% change in 5 sec = ~4% per min = significant

# Approximate free-float market cap (₹) per index — used for Traded Value %
# Update these periodically as market conditions change
MARKET_CAP_INR = {
    "NIFTY":     3.00e14,   # ≈ ₹300 lakh crore  (NSE total market cap)
    "BANKNIFTY": 4.50e13,   # ≈ ₹45  lakh crore  (banking sector subset)
    "SENSEX":    2.80e14,   # ≈ ₹280 lakh crore  (BSE total market cap)
}

# Approximate combined market cap of index constituent stocks only (₹)
# Used as denominator for "% of Index Market Cap" (equity TV column)
INDEX_EQ_MARKET_CAP_INR = {
    "NIFTY":     2.00e14,   # ≈ ₹200 lakh crore  (Nifty 50 stocks combined)
    "BANKNIFTY": 2.50e13,   # ≈ ₹25  lakh crore  (BankNifty 12 stocks)
    "SENSEX":    1.75e14,   # ≈ ₹175 lakh crore  (Sensex 30 stocks)
}

# NSE/BSE equity trading symbols of index constituents
# Used to subscribe equity tokens and compute index equity traded value
INDEX_CONSTITUENTS = {
    "NIFTY": [
        "RELIANCE",   "TCS",        "HDFCBANK",   "BHARTIARTL", "ICICIBANK",
        "INFY",       "SBIN",       "HINDUNILVR", "ITC",        "LT",
        "KOTAKBANK",  "AXISBANK",   "MARUTI",     "TITAN",      "SUNPHARMA",
        "BAJFINANCE", "HCLTECH",    "WIPRO",      "ADANIENT",   "ADANIPORTS",
        "ULTRACEMCO", "NTPC",       "POWERGRID",  "COALINDIA",  "ONGC",
        "JSWSTEEL",   "TATASTEEL",  "HINDALCO",   "TATAMOTORS", "M&M",
        "DRREDDY",    "CIPLA",      "DIVISLAB",   "APOLLOHOSP", "BAJAJFINSV",
        "BAJAJ-AUTO", "EICHERMOT",  "HEROMOTOCO", "TATACONSUM", "NESTLEIND",
        "BRITANNIA",  "GRASIM",     "SHRIRAMFIN", "BEL",        "INDUSINDBK",
        "TECHM",      "HDFCLIFE",   "SBILIFE",    "BPCL",       "ASIANPAINT",
    ],
    "BANKNIFTY": [
        "HDFCBANK",   "ICICIBANK",  "AXISBANK",   "KOTAKBANK",  "SBIN",
        "INDUSINDBK", "BANDHANBNK", "FEDERALBNK", "IDFCFIRSTB", "PNB",
        "BANKBARODA", "AUBANK",
    ],
    "SENSEX": [
        "RELIANCE",   "TCS",        "HDFCBANK",   "BHARTIARTL", "ICICIBANK",
        "INFY",       "SBIN",       "HINDUNILVR", "ITC",        "LT",
        "KOTAKBANK",  "AXISBANK",   "MARUTI",     "TITAN",      "SUNPHARMA",
        "BAJFINANCE", "HCLTECH",    "WIPRO",      "ADANIENT",   "NTPC",
        "POWERGRID",  "ULTRACEMCO", "TATASTEEL",  "TATAMOTORS", "M&M",
        "DRREDDY",    "BAJAJFINSV", "BAJAJ-AUTO", "TATACONSUM", "NESTLEIND",
    ],
}

# Rows reserved in OI Log before data
LOG_HDR_ROWS = 2         # row 1: title, row 2: column headers
LOG_DATA_ROW = 3         # first data row

# How many rows pre-formatted in the skeleton (covers a full trading day + buffer)
MAX_LOG_ROWS = 6000      # 375 min × 12 rows/min = 4500; buffer to 6000

# Chart data area in OI Chart sheet
CV_START      = 6        # first data row in chart view area (rows 1-5 = headers)
CV_MAX_ROWS   = 5100     # pre-allocated rows for chart series range

# ── Column indices in OI Log (1-based) ───────────────────────────────────────
CA, CB, CC, CD, CE_, CF, CG, CH, CI, CJ, CK, CL, CM, CN, CO, CP, CQ, CR, CS, CT = range(1, 21)
# A:ts  B:spot  C:atm  D:ce_itm  E:ce_otm  F:total_ce
# G:pe_otm  H:pe_itm  I:total_pe  J:pcr  K:net_bias
# L:oi_imbalance  M:dce  N:dpe  O:dpcr  P:signal
# Q:options_traded_value_5s  R:pct_of_total_mktcap
# S:equity_traded_value_5s   T:pct_of_index_mktcap   ← NEW


# ══════════════════════════════════════════════════════════════════════════════
# COLOUR / STYLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
NAV  = "0B2545"
GRN1 = "1E8449"; GRN2 = "EAFAF1"; GRN3 = "A9DFBF"
RED1 = "922B21"; RED2 = "FDEDEC"; RED3 = "F5B7B1"
BLU1 = "1565C0"; BLU2 = "AED6F1"; BLU3 = "D6EAF8"; BLU4 = "1F618D"
PRP1 = "6C3483"; PRP2 = "EAD7F7"
ORG1 = "B7770D"; ORG2 = "FEF9E7"; ORG3 = "FAD7A0"
GRY1 = "F2F3F4"; GRY2 = "BDC3C7"
WHT  = "FFFFFF"

_TH  = Side(style="thin",   color="CCCCCC")
_MED = Side(style="medium", color="888888")
BORD_TH  = Border(left=_TH,  right=_TH,  top=_TH,  bottom=_TH)
BORD_MED = Border(left=_MED, right=_MED, top=_MED, bottom=_MED)

def _f(h):             return PatternFill("solid", start_color=h, fgColor=h)
def _bfont(c, s=10):   return Font(bold=True,  color=c, size=s, name="Calibri")
def _rfont(c, s=10):   return Font(bold=False, color=c, size=s, name="Calibri")
def _ctr():            return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _lft():            return Alignment(horizontal="left",   vertical="center")
def _rgt():            return Alignment(horizontal="right",  vertical="center")
def _col_w(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

def _hdr(ws, row, col, val, bg, fg=WHT, bold=True, size=10):
    c = ws.cell(row, col, val)
    c.font = _bfont(fg, size) if bold else _rfont(fg, size)
    c.fill = _f(bg); c.alignment = _ctr(); c.border = BORD_MED
    return c

def _dat(ws, row, col, val, bg, fg="1A1A1A", align="right", fmt=None, size=9):
    c = ws.cell(row, col, val)
    c.font = _rfont(fg, size); c.fill = _f(bg)
    c.alignment = _lft() if align == "left" else _rgt()
    c.border = BORD_TH
    if fmt: c.number_format = fmt
    return c


# ══════════════════════════════════════════════════════════════════════════════
# GLOBAL STATE
# ══════════════════════════════════════════════════════════════════════════════
live_data   = {}
data_lock   = threading.Lock()
meta        = {}
running     = True
kws         = None
tick_in_sec = 0
ctr_lock    = threading.Lock()

# CSV backup — file handle + writer kept open for the session
csv_fh      = None
csv_writer  = None

# In-memory full log: each tuple is one row — all 16 fields
# (ts, spot, atm, ce_itm, ce_otm, tot_ce, pe_otm, pe_itm, tot_pe,
#  pcr, net_bias, oi_imb, dce, dpe, dpcr, signal)
oi_log: collections.deque = collections.deque(maxlen=CV_MAX_ROWS + 600)


# ══════════════════════════════════════════════════════════════════════════════
# WEBSOCKET CALLBACKS
# ══════════════════════════════════════════════════════════════════════════════
def on_ticks(ws_obj, ticks):
    global tick_in_sec
    with data_lock:
        for t in ticks:
            live_data[int(t["instrument_token"])] = t
    with ctr_lock:
        tick_in_sec += len(ticks)

def on_connect(ws_obj, response):
    toks = list(meta.get("all_toks", []))
    if toks:
        ws_obj.subscribe(toks)
        ws_obj.set_mode(ws_obj.MODE_FULL, toks)
    print(f"\n  ✓ WebSocket connected — {len(toks)} tokens subscribed")

def on_close(ws_obj, c, r):    print(f"\n  WS closed [{c}] {r}")
def on_error(ws_obj, c, r):    print(f"\n  WS error  [{c}] {r}")
def on_reconnect(ws_obj, n):   print(f"\n  Reconnecting #{n}")
def on_noreconnect(ws_obj):
    global running; running = False; print("\n  Max reconnects hit — stopping.")

def _tick(tok):
    with data_lock: return dict(live_data.get(tok) or {})


# ══════════════════════════════════════════════════════════════════════════════
# OI COMPUTATION + DERIVED METRICS
# ══════════════════════════════════════════════════════════════════════════════
def _safe_oi(tok_dict):
    return int(tok_dict.get("oi") or 0)

def compute_metrics(prev_row):
    """
    Compute all 16 metrics for one 5-second cycle.

    Boundary: spot (live float), NOT nearest ATM strike.
      CE ITM : k <  spot   (call has intrinsic value)
      CE OTM : k >= spot   (ATM included → no strike ever dropped)
      PE OTM : k <= spot   (ATM included → no strike ever dropped)
      PE ITM : k >  spot   (put has intrinsic value)

    Returns: dict with all fields + debug string.
    """
    strikes = meta["strikes"]
    ce_map  = meta["ce_map"]
    pe_map  = meta["pe_map"]
    spot    = meta["spot"]

    # ATM = nearest listed strike (display only, not the boundary)
    atm = min(strikes, key=lambda x: abs(x - spot))

    # ── Quadrant OI (accurate, every strike counted exactly once) ─────────────
    ce_itm = sum(_safe_oi(_tick(ce_map[k])) for k in strikes if k in ce_map and k <  spot)
    ce_otm = sum(_safe_oi(_tick(ce_map[k])) for k in strikes if k in ce_map and k >= spot)
    pe_otm = sum(_safe_oi(_tick(pe_map[k])) for k in strikes if k in pe_map and k <= spot)
    pe_itm = sum(_safe_oi(_tick(pe_map[k])) for k in strikes if k in pe_map and k >  spot)

    tot_ce = ce_itm + ce_otm
    tot_pe = pe_otm + pe_itm

    # ── Core ratios ───────────────────────────────────────────────────────────
    pcr = round(tot_pe / tot_ce, 4) if tot_ce > 0 else 0.0

    # Net Bias: CE_OTM − PE_OTM
    #   +ve → more call writing above spot (writers expect ceiling = bearish lean)
    #   -ve → more put writing below spot  (writers expect floor = bullish lean)
    net_bias = ce_otm - pe_otm

    # OI Imbalance: CE_ITM − PE_ITM
    #   +ve → more ITM CE than ITM PE (calls deeply held = bullish demand)
    #   -ve → more ITM PE held (put holders in profit = bearish demand)
    oi_imbalance = ce_itm - pe_itm

    # ── Momentum (requires prev row) ─────────────────────────────────────────
    if prev_row is not None:
        prev_tot_ce   = prev_row["tot_ce"]
        prev_tot_pe   = prev_row["tot_pe"]
        prev_pcr      = prev_row["pcr"]
        prev_spot     = prev_row["spot"]
        dce           = tot_ce - prev_tot_ce
        dpe           = tot_pe - prev_tot_pe
        dpcr          = round(pcr - prev_pcr, 4)
        dspot         = round(spot - prev_spot, 2)
        dce_pct       = dce / prev_tot_ce * 100 if prev_tot_ce > 0 else 0.0
        dpe_pct       = dpe / prev_tot_pe * 100 if prev_tot_pe > 0 else 0.0
        dce_otm_pct   = (ce_otm - prev_row["ce_otm"]) / prev_row["ce_otm"] * 100 if prev_row["ce_otm"] > 0 else 0.0
        dpe_otm_pct   = (pe_otm - prev_row["pe_otm"]) / prev_row["pe_otm"] * 100 if prev_row["pe_otm"] > 0 else 0.0
    else:
        dce = dpe = dpcr = dspot = 0
        dce_pct = dpe_pct = dce_otm_pct = dpe_otm_pct = 0.0

    # ── Algo Signal ───────────────────────────────────────────────────────────
    signal = _derive_signal(
        pcr, dce, dpe, dspot,
        dce_pct, dpe_pct,
        dce_otm_pct, dpe_otm_pct,
    )

    # ── Total Traded Value for the 5-second window ────────────────────────────
    # Uses cumulative volume from MODE_FULL ticks; delta per cycle ≈ 5s activity
    lot_sizes  = meta.get("lot_sizes", {})
    prev_vols  = meta.get("prev_volumes", {})
    curr_vols  = {}
    traded_val = 0.0

    for k in strikes:
        for tok_map in (ce_map, pe_map):
            if k in tok_map:
                tok  = tok_map[k]
                td   = _tick(tok)
                cv   = int(td.get("volume_traded") or td.get("volume") or 0)
                curr_vols[tok] = cv
                # delta_vol: contracts traded in last 5s (guard against resets)
                dv   = max(0, cv - prev_vols.get(tok, cv))
                lp   = float(td.get("last_price") or 0.0)
                lot  = lot_sizes.get(tok, 1)
                traded_val += dv * lp * lot

    meta["prev_volumes"] = curr_vols

    market_cap     = meta.get("market_cap", 1.0)
    traded_val_pct = traded_val / market_cap if market_cap > 0 else 0.0

    # ── Index Equity Traded Value for the 5-second window ─────────────────────
    # Tracks actual cash-market (equity) turnover of all constituent stocks
    # in the index (e.g. all 50 Nifty stocks), not options.
    # Formula: Σ [ delta_volume_in_5s × last_equity_price ]  (no lot size — qty is shares)
    eq_toks      = meta.get("eq_toks", {})          # {token: symbol}
    eq_prev_vols = meta.get("eq_prev_volumes", {})
    eq_curr_vols = {}
    eq_traded_val = 0.0

    for tok in eq_toks:
        td  = _tick(tok)
        cv  = int(td.get("volume_traded") or td.get("volume") or 0)
        eq_curr_vols[tok] = cv
        dv  = max(0, cv - eq_prev_vols.get(tok, cv))   # shares traded in last 5s
        lp  = float(td.get("last_price") or 0.0)
        eq_traded_val += dv * lp                        # ₹ = shares × price

    meta["eq_prev_volumes"] = eq_curr_vols

    eq_mktcap      = meta.get("eq_market_cap", 1.0)
    eq_traded_pct  = eq_traded_val / eq_mktcap if eq_mktcap > 0 else 0.0

    result = dict(
        spot=round(spot, 2), atm=int(atm),
        ce_itm=ce_itm, ce_otm=ce_otm, tot_ce=tot_ce,
        pe_otm=pe_otm, pe_itm=pe_itm, tot_pe=tot_pe,
        pcr=pcr, net_bias=net_bias, oi_imbalance=oi_imbalance,
        dce=dce, dpe=dpe, dpcr=dpcr, signal=signal,
        traded_val_5s=traded_val,
        traded_val_pct=traded_val_pct,
        eq_traded_val_5s=eq_traded_val,       # ← NEW
        eq_traded_pct=eq_traded_pct,           # ← NEW
    )

    # Accuracy cross-check string (terminal only)
    debug = (
        f"spot={spot:,.2f} ATM={int(atm):,} | "
        f"CE[{ce_itm:,}+{ce_otm:,}={tot_ce:,}] "
        f"PE[{pe_otm:,}+{pe_itm:,}={tot_pe:,}] | "
        f"PCR={pcr:.3f} ΔCE={dce:+,} ΔPE={dpe:+,} | {signal} | "
        f"OPT_TV=₹{traded_val:,.0f} ({traded_val_pct:.5%}) | "
        f"EQ_TV=₹{eq_traded_val:,.0f} ({eq_traded_pct:.5%})"
    )
    result["_debug"] = debug
    return result


def _derive_signal(pcr, dce, dpe, dspot,
                   dce_pct, dpe_pct, dce_otm_pct, dpe_otm_pct):
    """
    Deterministic signal from the 8 metrics above.
    Priority order: strongest conditions checked first.
    """
    surging = OI_SURGE_PCT   # threshold for "sharply rising"

    # ── Tier 1: STRONG signals (3-condition confluence) ──────────────────────
    if pcr > PCR_OVERSOLD and dce < 0 and dspot > 0:
        return "STRONG BUY"
    if pcr < PCR_OVERBOUGHT and dpe < 0 and dspot < 0:
        return "STRONG SELL"

    # ── Tier 2: Wall signals (OI piling at resistance/support) ───────────────
    if dce_otm_pct > surging and dspot > 0:
        return "SELL — RESISTANCE WALL"
    if dpe_otm_pct > surging and dspot < 0:
        return "BUY — SUPPORT FLOOR"

    # ── Tier 3: Buildup/Unwinding (2-condition) ───────────────────────────────
    if dce > 0 and dspot > 0:
        return "BULL CONFIRM — Long Buildup"
    if dce > 0 and dspot < 0:
        return "BEAR CONFIRM — Short Buildup"
    if dce < 0 and dspot > 0:
        return "COVER RALLY — Short Covering"
    if dpe < 0 and dspot < 0:
        return "UNWIND SELL — Long Unwinding"

    # ── Tier 4: PCR-only soft signals ────────────────────────────────────────
    if pcr > PCR_OVERSOLD:
        return "BULLISH BIAS — High PCR"
    if pcr < PCR_OVERBOUGHT:
        return "BEARISH BIAS — Low PCR"

    return "NEUTRAL"


# ══════════════════════════════════════════════════════════════════════════════
# BUILD — OI Log sheet skeleton
# ══════════════════════════════════════════════════════════════════════════════
_LOG_HEADERS = [
    # (col, label, bg, width)
    (CA,  "TIME\n(HH:MM:SS)",             NAV,  12),
    (CB,  "SPOT\nPRICE",                  NAV,  11),
    (CC,  "ATM\nSTRIKE",                  NAV,  11),
    (CD,  "CE ITM\n(k < spot)",           BLU1, 14),
    (CE_, "CE OTM\n(k ≥ spot)",           BLU4, 14),
    (CF,  "TOTAL\nCE OI",                 BLU1, 14),
    (CG,  "PE OTM\n(k ≤ spot)",           RED1, 14),
    (CH,  "PE ITM\n(k > spot)",           PRP1, 14),
    (CI,  "TOTAL\nPE OI",                 RED1, 14),
    (CJ,  "PCR\nPE/CE",                   ORG1, 10),
    (CK,  "NET BIAS\nCE_OTM−PE_OTM",      GRN1, 14),
    (CL,  "OI IMBAL.\nCE_ITM−PE_ITM",     GRN1, 14),
    (CM,  "ΔCE OI\n(vs prev)",            BLU1, 13),
    (CN,  "ΔPE OI\n(vs prev)",            RED1, 13),
    (CO,  "ΔPCR\n(vs prev)",              ORG1, 10),
    (CP,  "ALGO\nSIGNAL",                 NAV,  28),
    (CQ,  "TOTAL TRADED\nVALUE (5s ₹)",   BLU1, 18),
    (CR,  "% OF MKT\nCAP (5s)",           ORG1, 12),
    (CS,  "INDEX EQUITY\nTV (5s ₹)",      GRN1, 18),   # ← NEW: equity TV of index stocks
    (CT,  "% OF INDEX\nMKT CAP",          PRP1, 13),   # ← NEW: equity TV as % of index mktcap
]

# Number formats per column
_LOG_FMT = {
    CA: "@",         CB: "#,##0.00",   CC: "#,##0",
    CD: "#,##0",     CE_: "#,##0",     CF: "#,##0",
    CG: "#,##0",     CH: "#,##0",      CI: "#,##0",
    CJ: "0.000",     CK: "#,##0",      CL: "#,##0",
    CM: "+#,##0;-#,##0;0",
    CN: "+#,##0;-#,##0;0",
    CO: "+0.000;-0.000;0.000",
    CP: "@",
    CQ: "₹#,##0",           # Total Options Traded Value (₹)
    CR: "0.00000%",          # % of Total Market Cap (5 decimal places)
    CS: "₹#,##0",           # Index Equity Traded Value (₹)
    CT: "0.00000%",          # % of Index Market Cap (5 decimal places)
}

# Alternating row colours per column type
_COL_BG = {
    CA: ("EAECEE", "F2F3F4"),
    CB: ("EBF5FB", "D6EAF8"),
    CC: ("EBF5FB", "D6EAF8"),
    CD: ("D6EAF8", "EBF5FB"),
    CE_:("AED6F1", "D6EAF8"),
    CF: ("AED6F1", "D6EAF8"),
    CG: ("FADBD8", "FDEBD0"),
    CH: ("EAD7F7", "F9EBFF"),
    CI: ("FADBD8", "FDEBD0"),
    CJ: ("FEF9E7", "FEF5E7"),
    CK: ("EAFAF1", "D5F5E3"),
    CL: ("EAFAF1", "D5F5E3"),
    CM: ("EBF5FB", "D6EAF8"),
    CN: ("FADBD8", "FDEBD0"),
    CO: ("FEF9E7", "FEF5E7"),
    CP: ("F2F3F4", "EAECEE"),
    CQ: ("D6EAF8", "EBF5FB"),   # Options Traded Value — blue tones
    CR: ("FEF9E7", "FEF5E7"),   # % Total Mkt Cap      — orange tones
    CS: ("EAFAF1", "D5F5E3"),   # Equity Traded Value  — green tones   ← NEW
    CT: ("EAD7F7", "F9EBFF"),   # % Index Mkt Cap      — purple tones  ← NEW
}


def build_log_sheet(ws):
    # ── Row 1: title banner ───────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(CT)}1")
    c = ws.cell(1, 1, "  PROFESSIONAL OI ANALYSIS LOG  —  5-Second Tick Record  "
                       "|  PCR · Net Bias · OI Momentum · Algo Signal · Options TV · Index Equity TV  |  Accuracy Verified")
    c.font = _bfont(WHT, 11); c.fill = _f(NAV); c.alignment = _lft(); c.border = BORD_MED
    ws.row_dimensions[1].height = 26

    # ── Row 2: column headers ────────────────────────────────────────────────
    for col, label, bg, width in _LOG_HEADERS:
        _hdr(ws, 2, col, label, bg, WHT, bold=True, size=9)
        _col_w(ws, col, width)
    ws.row_dimensions[2].height = 34

    # ── Pre-format data rows (number format + alternating colours) ────────────
    for ri in range(MAX_LOG_ROWS):
        row = LOG_DATA_ROW + ri
        parity = ri % 2
        for col, _, _, _ in _LOG_HEADERS:
            c = ws.cell(row, col)
            c.number_format = _LOG_FMT.get(col, "General")
            bg0, bg1 = _COL_BG.get(col, ("FFFFFF", "F8F8F8"))
            c.fill      = _f(bg0 if parity == 0 else bg1)
            c.font      = _rfont("1A1A1A", 9)
            c.alignment = _lft() if col in (CA, CP) else _rgt()
            c.border    = BORD_TH
        ws.row_dimensions[row].height = 15

    # ── Conditional formatting on Signal column (P) ───────────────────────────
    sig_range = f"P{LOG_DATA_ROW}:P{LOG_DATA_ROW + MAX_LOG_ROWS - 1}"
    # Green fill for bullish signals
    for kw in ("STRONG BUY", "BUY", "BULL CONFIRM", "BULLISH", "COVER RALLY", "SUPPORT FLOOR"):
        ws.conditional_formatting.add(
            sig_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{kw}",P{LOG_DATA_ROW}))'],
                fill=_f(GRN2), font=_bfont(GRN1, 9)
            )
        )
    # Red fill for bearish signals
    for kw in ("STRONG SELL", "SELL", "BEAR CONFIRM", "BEARISH", "UNWIND SELL", "RESISTANCE"):
        ws.conditional_formatting.add(
            sig_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("{kw}",P{LOG_DATA_ROW}))'],
                fill=_f(RED2), font=_bfont(RED1, 9)
            )
        )

    # ── PCR colour scale: green (high) → red (low) ───────────────────────────
    # (openpyxl ColorScaleRule applied to column J)
    from openpyxl.formatting.rule import ColorScaleRule
    pcr_range = f"J{LOG_DATA_ROW}:J{LOG_DATA_ROW + MAX_LOG_ROWS - 1}"
    ws.conditional_formatting.add(
        pcr_range,
        ColorScaleRule(
            start_type="num", start_value=0.5, start_color="E74C3C",
            mid_type="num",   mid_value=1.0,   mid_color="F39C12",
            end_type="num",   end_value=1.5,   end_color="27AE60",
        )
    )

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = f"A{LOG_DATA_ROW}"


# ══════════════════════════════════════════════════════════════════════════════
# BUILD — OI Chart sheet (3 charts, no timeframe dropdown)
# ══════════════════════════════════════════════════════════════════════════════

# Column layout in OI Chart sheet:
# A: timestamp   B: ce_itm   C: ce_otm   D: pe_otm   E: pe_itm
# F: tot_ce      G: tot_pe   H: pcr      I: net_bias  J: dce   K: dpe

_CHT_HEADERS = [
    (1, "TIME",        NAV),
    (2, "CE ITM",      BLU1),
    (3, "CE OTM",      BLU4),
    (4, "PE OTM",      RED1),
    (5, "PE ITM",      PRP1),
    (6, "TOTAL CE",    BLU1),
    (7, "TOTAL PE",    RED1),
    (8, "PCR",         ORG1),
    (9, "NET BIAS",    GRN1),
    (10, "ΔCE OI",    BLU1),
    (11, "ΔPE OI",    RED1),
    (12, "TRADED VAL\n(5s ₹)", BLU1),
    (13, "% MKT\nCAP", ORG1),
]
_CHT_FMT = {1:"@", 2:"#,##0", 3:"#,##0", 4:"#,##0", 5:"#,##0",
            6:"#,##0", 7:"#,##0", 8:"0.000", 9:"+#,##0;-#,##0;0",
            10:"+#,##0;-#,##0;0", 11:"+#,##0;-#,##0;0",
            12:"₹#,##0", 13:"0.00000%"}


def build_chart_sheet(ws):
    # ── Row 1: title ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    c = ws.cell(1, 1, "  OI ANALYSIS  —  3 LIVE CHARTS  "
                       "|  Quadrant Trend  ·  PCR Signal  ·  OI Build Rate  ·  Traded Value  |  Full Day View")
    c.font = _bfont(WHT, 12); c.fill = _f(NAV); c.alignment = _lft(); c.border = BORD_MED
    ws.row_dimensions[1].height = 28

    # ── Row 2: chart legend row ───────────────────────────────────────────────
    legend = [
        (2,  BLU2, "CE ITM (< spot)"),
        (3,  BLU3, "CE OTM (≥ spot)"),
        (4,  RED2, "PE OTM (≤ spot)"),
        (5,  PRP2, "PE ITM (> spot)"),
        (6,  BLU2, "Total CE OI"),
        (7,  RED2, "Total PE OI"),
        (8,  ORG2, "PCR  =  PE / CE"),
        (9,  GRN2, "Net Bias  CE_OTM−PE_OTM"),
        (10, BLU3, "ΔCE OI / 5s"),
        (11, RED3, "ΔPE OI / 5s"),
        (12, BLU2, "Traded Val (5s ₹)"),
        (13, ORG2, "% Mkt Cap"),
    ]
    ws.merge_cells("A2:A2")
    c = ws.cell(2, 1, "LEGEND  ▶")
    c.font = _bfont(WHT, 9); c.fill = _f(NAV); c.alignment = _ctr(); c.border = BORD_MED
    for col, bg, lbl in legend:
        c = ws.cell(2, col, lbl)
        c.font = _bfont("1A1A1A", 9); c.fill = _f(bg)
        c.alignment = _ctr(); c.border = BORD_MED
    ws.row_dimensions[2].height = 20

    # ── Row 3: PCR threshold note ─────────────────────────────────────────────
    ws.merge_cells("A3:M3")
    note = (f"  ℹ  PCR > {PCR_OVERSOLD} = Oversold (bullish zone)  |  "
            f"PCR < {PCR_OVERBOUGHT} = Overbought (bearish zone)  |  "
            f"Net Bias +ve = CE writing heavy above spot (resistance)  |  "
            f"ΔCE > 0 & spot ↑ = Long Buildup  |  "
            f"ΔCE > 0 & spot ↓ = Short Buildup")
    c = ws.cell(3, 1, note)
    c.font = Font(name="Calibri", italic=True, size=8, color="444444")
    c.fill = _f("F7F9FA"); c.alignment = _lft(); c.border = BORD_TH
    ws.row_dimensions[3].height = 16

    # ── Row 4: signal summary (updated live) ─────────────────────────────────
    ws.merge_cells("A4:M4")
    c = ws.cell(4, 1, "CURRENT SIGNAL  →  (updating...)")
    c.font = _bfont("1A1A1A", 10); c.fill = _f(GRY1)
    c.alignment = _lft(); c.border = BORD_MED
    ws.row_dimensions[4].height = 22

    # ── Row 5: chart-data column headers ─────────────────────────────────────
    for col, label, bg in _CHT_HEADERS:
        _hdr(ws, 5, col, label, bg, WHT, bold=True, size=9)
        _col_w(ws, col, 14)
    ws.row_dimensions[5].height = 20

    # ── Rows CV_START … : pre-format chart data area ──────────────────────────
    alt_bg = ["EBF5FB", "D6EAF8"]
    for ri in range(CV_MAX_ROWS):
        row = CV_START + ri
        bg  = alt_bg[ri % 2]
        for col, _, _ in _CHT_HEADERS:
            c = ws.cell(row, col)
            c.number_format = _CHT_FMT.get(col, "General")
            c.fill = _f(bg); c.font = _rfont("1A1A1A", 9)
            c.alignment = _lft() if col == 1 else _rgt()
            c.border = BORD_TH
        ws.row_dimensions[row].height = 14

    ws.freeze_panes = f"A{CV_START}"

    # ── 3 embedded charts ─────────────────────────────────────────────────────
    _add_chart1_quadrant_trend(ws)
    _add_chart2_pcr_bias(ws)
    _add_chart3_oi_delta(ws)

    # Column widths
    _col_w(ws, 1, 11)
    for ci in range(2, 14):
        _col_w(ws, ci, 13)


def _add_chart1_quadrant_trend(ws):
    """Line chart: 4 OI quadrants over time."""
    chart = LineChart()
    chart.title  = "OI Quadrant Trend — CE ITM / CE OTM / PE OTM / PE ITM"
    chart.style  = 10
    chart.y_axis.title = "Open Interest (contracts)"
    chart.x_axis.title = "Time →"
    chart.width  = 26; chart.height = 14
    chart.grouping = "standard"
    chart.y_axis.numFmt = "#,##0"

    series_cfg = [
        (2, BLU1, "CE ITM  (below spot)"),
        (3, BLU4, "CE OTM  (above spot)"),
        (4, RED1, "PE OTM  (below spot)"),
        (5, PRP1, "PE ITM  (above spot)"),
    ]
    for col, color, label in series_cfg:
        dr = Reference(ws, min_col=col, max_col=col,
                       min_row=CV_START, max_row=CV_START + CV_MAX_ROWS - 1)
        chart.add_data(dr)
        s = chart.series[-1]
        s.title = SeriesLabel(v=label)
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width     = 15000
        s.smooth = True

    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=CV_START, max_row=CV_START + CV_MAX_ROWS - 1)
    chart.set_categories(cats)
    ws.add_chart(chart, "M2")          # anchor: right of data columns


def _add_chart2_pcr_bias(ws):
    """Line chart: PCR trend + Net Bias combo."""
    # PCR line
    chart = LineChart()
    chart.title  = "PCR Trend + Net OI Bias  |  PCR > 1.3 = Bullish Zone  |  PCR < 0.7 = Bearish Zone"
    chart.style  = 10
    chart.y_axis.title  = "PCR value"
    chart.x_axis.title  = "Time →"
    chart.width  = 26; chart.height = 14
    chart.grouping = "standard"
    chart.y_axis.numFmt = "0.000"

    # PCR
    dr_pcr = Reference(ws, min_col=8, max_col=8,
                        min_row=CV_START, max_row=CV_START + CV_MAX_ROWS - 1)
    chart.add_data(dr_pcr)
    s = chart.series[-1]
    s.title = SeriesLabel(v="PCR  (PE / CE)")
    s.graphicalProperties.line.solidFill = ORG1
    s.graphicalProperties.line.width     = 20000
    s.smooth = True

    # Net Bias on secondary axis
    dr_bias = Reference(ws, min_col=9, max_col=9,
                         min_row=CV_START, max_row=CV_START + CV_MAX_ROWS - 1)
    chart.add_data(dr_bias)
    s2 = chart.series[-1]
    s2.title = SeriesLabel(v="Net Bias  CE_OTM − PE_OTM")
    s2.graphicalProperties.line.solidFill = GRN1
    s2.graphicalProperties.line.width     = 12000
    s2.smooth = True

    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=CV_START, max_row=CV_START + CV_MAX_ROWS - 1)
    chart.set_categories(cats)
    ws.add_chart(chart, "M32")         # below Chart 1


def _add_chart3_oi_delta(ws):
    """Bar chart: ΔCE and ΔPE per 5-sec cycle."""
    chart = BarChart()
    chart.type     = "col"
    chart.grouping = "clustered"
    chart.title    = "OI Build Rate  —  ΔCE OI / ΔPE OI per 5-sec tick  (+ = buildup  − = unwinding)"
    chart.style    = 10
    chart.y_axis.title  = "OI Change (contracts)"
    chart.x_axis.title  = "Time →"
    chart.width    = 26; chart.height = 14
    chart.y_axis.numFmt = "+#,##0;-#,##0;0"

    series_cfg = [
        (10, BLU1, "ΔCE OI"),
        (11, RED1, "ΔPE OI"),
    ]
    for col, color, label in series_cfg:
        dr = Reference(ws, min_col=col, max_col=col,
                       min_row=CV_START, max_row=CV_START + CV_MAX_ROWS - 1)
        chart.add_data(dr)
        s = chart.series[-1]
        s.title = SeriesLabel(v=label)
        s.graphicalProperties.solidFill = color

    cats = Reference(ws, min_col=1, max_col=1,
                     min_row=CV_START, max_row=CV_START + CV_MAX_ROWS - 1)
    chart.set_categories(cats)
    ws.add_chart(chart, "M62")         # below Chart 2


# ══════════════════════════════════════════════════════════════════════════════
# LIVE PATCH — write one row to log + refresh chart data area
# ══════════════════════════════════════════════════════════════════════════════
def patch_log(xws_log, ts_str, m, log_row):
    """Write one data row to OI Log sheet (20 values)."""
    row_vals = [[
        ts_str,
        m["spot"],   m["atm"],
        m["ce_itm"], m["ce_otm"], m["tot_ce"],
        m["pe_otm"], m["pe_itm"], m["tot_pe"],
        m["pcr"],    m["net_bias"], m["oi_imbalance"],
        m["dce"],    m["dpe"],    m["dpcr"],
        m["signal"],
        m["traded_val_5s"],
        m["traded_val_pct"],
        m["eq_traded_val_5s"],     # ← NEW col S
        m["eq_traded_pct"],        # ← NEW col T
    ]]
    xws_log.range(f"A{log_row}:{get_column_letter(CT)}{log_row}").value = row_vals


def patch_chart_view(xws_chart, signal):
    """
    Overwrite entire chart-view data area from the in-memory oi_log.
    Also update the live signal banner in row 4.
    """
    snapshot = list(oi_log)
    n = len(snapshot)
    if n == 0:
        return

    # Build 13-column matrix from log tuples
    # Log tuple: (ts, spot, atm, ce_itm, ce_otm, tot_ce, pe_otm, pe_itm, tot_pe,
    #              pcr, net_bias, oi_imb, dce, dpe, dpcr, signal,
    #              traded_val_5s, traded_val_pct,
    #              eq_traded_val_5s, eq_traded_pct)   ← indices 18,19
    values = [
        [
            r[0],   # A: timestamp
            r[3],   # B: ce_itm
            r[4],   # C: ce_otm
            r[6],   # D: pe_otm
            r[7],   # E: pe_itm
            r[5],   # F: tot_ce
            r[8],   # G: tot_pe
            r[9],   # H: pcr
            r[10],  # I: net_bias
            r[12],  # J: dce
            r[13],  # K: dpe
            r[16],  # L: traded_val_5s
            r[17],  # M: traded_val_pct
        ]
        for r in snapshot
    ]

    used_end = CV_START + n - 1
    xws_chart.range(f"A{CV_START}:M{used_end}").value = values

    # Clear unused rows
    tail_start = CV_START + n
    tail_end   = CV_START + CV_MAX_ROWS - 1
    if tail_start <= tail_end:
        clear_n = tail_end - tail_start + 1
        xws_chart.range(
            f"A{tail_start}:M{tail_end}"
        ).value = [[None] * 13] * clear_n

    # Signal banner (row 4)
    _sig_col = {
        "STRONG BUY":   ("STRONG BUY  ▲▲", GRN1),
        "BUY":          ("BUY  ▲",          GRN1),
        "BULL CONFIRM": ("BULL CONFIRM  ▲", "117A65"),
        "COVER RALLY":  ("COVER RALLY  ↑",  "1A5276"),
        "BULLISH":      ("BULLISH BIAS  ↗", "1D8348"),
        "STRONG SELL":  ("STRONG SELL  ▼▼", RED1),
        "SELL":         ("SELL  ▼",         RED1),
        "BEAR CONFIRM": ("BEAR CONFIRM  ▼", "7B241C"),
        "UNWIND SELL":  ("UNWIND SELL  ↓",  "922B21"),
        "BEARISH":      ("BEARISH BIAS  ↘", "7D3C98"),
    }
    disp, col_hex = next(
        ((v, c) for k, (v, c) in _sig_col.items() if k in signal.upper()),
        (f"  {signal}", "555555")
    )
    last = snapshot[-1]
    banner = (
        f"  LIVE SIGNAL  →  {disp}"
        f"    |    PCR: {last[9]:.3f}"
        f"    |    Spot: {last[1]:,.2f}"
        f"    |    ATM: {last[2]:,}"
        f"    |    ΔCE: {last[12]:+,}"
        f"    |    ΔPE: {last[13]:+,}"
        f"    |    Net Bias: {last[10]:+,}"
        f"    |    OPT TV(5s): ₹{last[16]:,.0f}"
        f"    |    EQ TV(5s): ₹{last[18]:,.0f}"
        f"    |    EQ TV%: {last[19]:.5%}"
    )
    xws_chart.range("A4:M4").merge()
    cell = xws_chart.range("A4")
    cell.value = banner
    cell.api.Font.Bold    = True
    cell.api.Font.Size    = 10
    cell.api.Font.Color   = int(col_hex, 16)
    # Background based on signal type
    is_bull = any(k in signal.upper() for k in ("BUY","BULL","COVER","BULLISH"))
    is_bear = any(k in signal.upper() for k in ("SELL","BEAR","UNWIND","BEARISH"))
    cell.color = (0xEA, 0xFA, 0xF1) if is_bull else (0xFD, 0xED, 0xEC) if is_bear else (0xF2, 0xF3, 0xF4)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    global running, kws, tick_in_sec

    print("\n" + "═" * 80)
    print("  Professional OI Analysis Recorder  —  5-Second Tick  |  16 Metrics  |  3 Charts")
    print("═" * 80)

    # ── Read zerodha_config.xlsx ──────────────────────────────────────────────
    print(f"\n  Reading {CONFIG_FILE} ...")
    try:
        from openpyxl import load_workbook as _lw
        _wb = _lw(CONFIG_FILE, read_only=True, data_only=True)
        _ws = _wb["Config"]
        _kv = {}; _sr = {}; _ex = {}
        _in_sr = _in_ex = False
        for _r in _ws.iter_rows(values_only=True):
            _a  = str(_r[0]).strip() if _r[0] else ""
            _b_ = _r[1] if len(_r) > 1 else None
            if _a == "Index" and _b_ == "Range (pts)":  _in_sr = True; _in_ex = False; continue
            if _a == "Index" and _b_ == "Expiry Date":  _in_ex = True; _in_sr = False; continue
            if _in_sr and _a in ("NIFTY","BANKNIFTY","SENSEX") and _b_ is not None: _sr[_a] = int(_b_)
            elif _in_ex and _a in ("NIFTY","BANKNIFTY","SENSEX") and _b_ is not None: _ex[_a] = str(_b_)
            elif _a and _b_ is not None and _a not in ("Index",): _kv[_a] = _b_
        _wb.close()
    except FileNotFoundError:
        sys.exit(f"\n  ERROR: {CONFIG_FILE} not found.\n")
    except Exception as e:
        sys.exit(f"\n  ERROR reading config: {e}\n")

    api_key      = str(_kv.get("api_key",       ""))
    access_token = str(_kv.get("access_token",  ""))
    rfr          = float(_kv.get("risk_free_rate", 0.065))
    opt_index    = str(_kv.get("options_index", "NIFTY")).upper()

    OPT_CFG = {
        "NIFTY":     {"exchange":"NFO", "spot_sym":"NSE:NIFTY 50",   "step":50,  "range":2500},
        "BANKNIFTY": {"exchange":"NFO", "spot_sym":"NSE:NIFTY BANK", "step":100, "range":2500},
        "SENSEX":    {"exchange":"BFO", "spot_sym":"BSE:SENSEX",     "step":100, "range":1000},
    }
    if opt_index not in OPT_CFG:
        sys.exit(f"\n  ERROR: options_index must be NIFTY / BANKNIFTY / SENSEX\n")

    ocfg         = OPT_CFG[opt_index]
    strike_range = _sr.get(opt_index, ocfg["range"])
    expiry_str   = _ex.get(opt_index, "")
    print(f"  ✓ Config  —  Index:{opt_index}  Expiry:{expiry_str}")

    # ── Authenticate ──────────────────────────────────────────────────────────
    kite = KiteConnect(api_key=api_key)
    kite.set_access_token(access_token)
    print("  ✓ Authenticated")

    # ── Spot price ────────────────────────────────────────────────────────────
    raw        = kite.quote([ocfg["spot_sym"]])
    spot_price = float(raw[ocfg["spot_sym"]]["last_price"])
    spot_tok   = int(raw[ocfg["spot_sym"]]["instrument_token"])
    print(f"  ✓ {opt_index} spot = {spot_price:,.2f}")

    # ── Expiry date ───────────────────────────────────────────────────────────
    try:
        expiry_date = datetime.date.fromisoformat(expiry_str)
    except Exception:
        today = datetime.date.today()
        expiry_date = today + datetime.timedelta((3 - today.weekday()) % 7)
    T = max(0.0, (expiry_date - datetime.date.today()).days / 365.0)

    # ── Option chain ─────────────────────────────────────────────────────────
    exch = ocfg["exchange"]
    print(f"  Downloading {exch} instruments ...")
    instruments = kite.instruments(exch)
    atm_approx  = round(spot_price / ocfg["step"]) * ocfg["step"]
    lo_s = atm_approx - strike_range
    hi_s = atm_approx + strike_range

    ce_map = {}; pe_map = {}; lot_sizes = {}
    for inst in instruments:
        if (inst["name"] == opt_index and inst["expiry"] == expiry_date
                and inst["instrument_type"] in ("CE","PE")
                and lo_s <= inst["strike"] <= hi_s):
            tok = int(inst["instrument_token"])
            lot_sizes[tok] = int(inst.get("lot_size", 1) or 1)
            if inst["instrument_type"] == "CE": ce_map[inst["strike"]] = tok
            else:                                pe_map[inst["strike"]] = tok

    strikes = sorted(set(ce_map) | set(pe_map))
    if not strikes:
        sys.exit("\n  ERROR: No options found — check expiry date.\n")

    atm = min(strikes, key=lambda x: abs(x - spot_price))
    print(f"  ✓ {len(strikes)} strikes  ATM={int(atm):,}  [{int(min(strikes)):,}–{int(max(strikes)):,}]")

    # ── Equity constituent tokens ─────────────────────────────────────────────
    # Download NSE equity instruments, match index constituent symbols,
    # subscribe their tokens so we get live volume + LTP every tick.
    eq_syms     = INDEX_CONSTITUENTS.get(opt_index, [])
    eq_exchange = "NSE"   # use NSE equity for all indices (incl. SENSEX — dual-listed)
    print(f"  Downloading {eq_exchange} equity instruments for {opt_index} constituents ...")
    try:
        eq_instruments = kite.instruments(eq_exchange)
    except Exception as _e:
        eq_instruments = []
        print(f"  [!] Could not fetch equity instruments: {_e}")

    eq_sym_map = {
        inst["tradingsymbol"]: inst
        for inst in eq_instruments
        if str(inst.get("segment", "")).startswith("NSE")   # covers "NSE" and "NSE-EQ"
        and inst.get("instrument_type") == "EQ"
    }
    eq_toks = {}          # {token(int): symbol(str)}
    missing = []
    for sym in eq_syms:
        inst = eq_sym_map.get(sym)
        if inst:
            eq_toks[int(inst["instrument_token"])] = sym
        else:
            missing.append(sym)
    if missing:
        print(f"  [!] Equity tokens not found (check symbols): {missing}")
    print(f"  ✓ {len(eq_toks)}/{len(eq_syms)} equity tokens resolved for {opt_index}")

    # ── Populate meta ─────────────────────────────────────────────────────────
    all_toks = [spot_tok] + list(ce_map.values()) + list(pe_map.values()) + list(eq_toks.keys())
    meta.update(spot=spot_price, spot_tok=spot_tok,
                strikes=strikes, ce_map=ce_map, pe_map=pe_map,
                T=T, rfr=rfr, opt_index=opt_index, expiry_date=expiry_date,
                all_toks=all_toks,
                lot_sizes=lot_sizes, prev_volumes={},
                market_cap=MARKET_CAP_INR.get(opt_index, 1e14),
                eq_toks=eq_toks,                                          # ← NEW
                eq_prev_volumes={},                                        # ← NEW
                eq_market_cap=INDEX_EQ_MARKET_CAP_INR.get(opt_index, 1e14))  # ← NEW
    with data_lock:
        live_data[spot_tok] = {"last_price": spot_price}
        for tok in list(ce_map.values()) + list(pe_map.values()):
            live_data[tok] = {}
        for tok in eq_toks:                  # ← NEW: init equity token slots
            live_data[tok] = {}

    # ── Build Excel workbook ──────────────────────────────────────────────────
    print(f"\n  Building workbook  →  {OUTPUT_FILE}")
    wb = Workbook(); wb.remove(wb.active)
    ws_log   = wb.create_sheet(SHEET_LOG)
    ws_chart = wb.create_sheet(SHEET_CHART)
    build_log_sheet(ws_log)
    build_chart_sheet(ws_chart)

    # Close any existing open copy
    fname = os.path.basename(OUTPUT_FILE)
    try:
        for app in xw.apps:
            for book in app.books:
                if book.name.lower() == fname.lower():
                    book.close()
    except Exception:
        pass

    try:
        wb.save(OUTPUT_FILE)
    except PermissionError:
        sys.exit(f"\n  ERROR: '{fname}' is open — close it and retry.\n")
    print("  ✓ Workbook saved")

    # ── Open CSV backup (append mode — survives mid-session restart) ──────────
    global csv_fh, csv_writer
    _csv_is_new = not os.path.exists(CSV_BACKUP_FILE)
    csv_fh     = open(CSV_BACKUP_FILE, "a", newline="", encoding="utf-8", buffering=1)
    csv_writer = csv.writer(csv_fh)
    if _csv_is_new:
        csv_writer.writerow([
            "TIME", "SPOT", "ATM",
            "CE_ITM", "CE_OTM", "TOTAL_CE",
            "PE_OTM", "PE_ITM", "TOTAL_PE",
            "PCR", "NET_BIAS", "OI_IMBALANCE",
            "DELTA_CE", "DELTA_PE", "DELTA_PCR",
            "SIGNAL", "OPT_TRADED_VAL_5S", "OPT_TRADED_VAL_PCT",
            "EQ_TRADED_VAL_5S", "EQ_TRADED_VAL_PCT",   # ← NEW
        ])
        csv_fh.flush()
    print(f"  ✓ CSV backup  →  {CSV_BACKUP_FILE}")

    # ── Open in Excel ─────────────────────────────────────────────────────────
    print("  Opening in Excel ...")
    try:
        xw_wb     = xw.Book(OUTPUT_FILE)
        xws_log   = xw_wb.sheets[SHEET_LOG]
        xws_chart = xw_wb.sheets[SHEET_CHART]
    except Exception as e:
        sys.exit(f"\n  ERROR: {e}\n  Ensure Microsoft Excel is installed.\n")

    # Apply column number formats via xlwings (more reliable than openpyxl on COM)
    for col, _, _, _ in _LOG_HEADERS:
        col_ltr = get_column_letter(col)
        fmt = _LOG_FMT.get(col, "General")
    xws_log.range(
            f"{col_ltr}{LOG_DATA_ROW}:{col_ltr}{LOG_DATA_ROW+MAX_LOG_ROWS-1}"
        ).number_format = fmt

    for col, _, _ in _CHT_HEADERS:
        col_ltr = get_column_letter(col)
        fmt = _CHT_FMT.get(col, "General")
        xws_chart.range(
            f"{col_ltr}{CV_START}:{col_ltr}{CV_START+CV_MAX_ROWS-1}"
        ).number_format = fmt

    print("  ✓ Excel open — 2 sheets, 3 charts ready")
    print(f"  ✓ Equity TV tracking: {len(eq_toks)} {opt_index} constituent stocks")

    # ── KiteTicker ───────────────────────────────────────────────────────────
    kws = KiteTicker(api_key, access_token)
    kws.on_ticks      = on_ticks
    kws.on_connect    = on_connect
    kws.on_close      = on_close
    kws.on_error      = on_error
    kws.on_reconnect  = on_reconnect
    kws.on_noreconnect = on_noreconnect
    print("  Connecting WebSocket ...")
    kws.connect(threaded=True)

    # ── Graceful shutdown ─────────────────────────────────────────────────────
    def _stop(sig_num, frame):
        global running; running = False
        print("\n\n  Stopping ...")
        if kws:
            try: kws.stop()
            except: pass
        try:
            xw_wb.save()
            print(f"  ✓ Saved  ({len(oi_log):,} rows)")
        except Exception:
            pass
        if csv_fh is not None:
            try: csv_fh.flush(); csv_fh.close()
            except: pass
            print(f"  ✓ CSV backup closed  →  {CSV_BACKUP_FILE}")
        print("  ✓ Excel left open.\n")
        sys.exit(0)

    signal.signal(signal.SIGINT,  _stop)
    signal.signal(signal.SIGTERM, _stop)

    # ── Main loop ─────────────────────────────────────────────────────────────
    log_row  = LOG_DATA_ROW
    cycle    = 0
    prev_row = None
    print(f"\n  Recording every {RECORD_SEC}s  (Ctrl+C to stop)\n  {'─'*80}")

    while running:
        t0 = time.time()
        cycle += 1

        # Refresh live spot from WebSocket tick
        sp_tick = _tick(spot_tok)
        if sp_tick.get("last_price"):
            meta["spot"] = float(sp_tick["last_price"])

        # Refresh time-to-expiry
        meta["T"] = max(0.0, (expiry_date - datetime.date.today()).days / 365.0)

        # Compute all 16 metrics
        ts_str = datetime.datetime.now().strftime("%H:%M:%S")
        m = compute_metrics(prev_row)

        # Store full tuple in in-memory log
        oi_log.append((
            ts_str,
            m["spot"], m["atm"],
            m["ce_itm"], m["ce_otm"], m["tot_ce"],
            m["pe_otm"], m["pe_itm"], m["tot_pe"],
            m["pcr"], m["net_bias"], m["oi_imbalance"],
            m["dce"], m["dpe"], m["dpcr"],
            m["signal"],
            m["traded_val_5s"],
            m["traded_val_pct"],
            m["eq_traded_val_5s"],    # index 18 ← NEW
            m["eq_traded_pct"],       # index 19 ← NEW
        ))

        # ── CSV backup: one row per cycle, flushed immediately ────────────────
        if csv_writer is not None:
            csv_writer.writerow(oi_log[-1])
            csv_fh.flush()

        # Write to Excel
        try:
            patch_log(xws_log, ts_str, m, log_row)
            patch_chart_view(xws_chart, m["signal"])
            log_row += 1
        except Exception as e:
            print(f"\r  [!] Excel write error: {e}  ", end="")

        # ── Auto-save Excel every AUTO_SAVE_CYCLES cycles ─────────────────────
        if cycle % AUTO_SAVE_CYCLES == 0:
            try:
                xw_wb.save()
                print(f"\r  [✓] Auto-saved Excel  (cycle {cycle}, {len(oi_log):,} rows)  ",
                      end="", flush=True)
            except Exception as _e:
                print(f"\r  [!] Auto-save failed: {_e}  ", end="")

        prev_row = m

        # Ticks/sec counter
        with ctr_lock:
            tps = tick_in_sec; tick_in_sec = 0

        # Terminal status — full cross-check every cycle
        print(f"\r  #{cycle:>5}  {ts_str}  │  {m['_debug']}  │  T/s:{tps}  Rows:{len(oi_log)}  ",
              end="", flush=True)

        elapsed = time.time() - t0
        time.sleep(max(0.0, RECORD_SEC - elapsed))


if __name__ == "__main__":
    main()
